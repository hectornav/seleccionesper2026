"""
Importa datos de transparencia ONPE desde reporte_if.xlsx.
Actualiza los campos transparencia_onpe_* de cada Partido.

Uso: python manage.py importar_onpe
     python manage.py importar_onpe --archivo /ruta/a/reporte_if.xlsx
"""
import unicodedata
from datetime import date
from pathlib import Path

from django.core.management.base import BaseCommand
from candidatos.models import Partido


def normalize(name):
    name = unicodedata.normalize('NFD', name)
    name = ''.join(c for c in name if unicodedata.category(c) != 'Mn')
    return ' '.join(name.upper().split())


class Command(BaseCommand):
    help = 'Importa datos de transparencia ONPE desde reporte_if.xlsx'

    def add_arguments(self, parser):
        parser.add_argument(
            '--archivo',
            default='reporte_if.xlsx',
            help='Ruta al archivo Excel de ONPE',
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stderr.write(self.style.ERROR('Necesitas openpyxl: pip install openpyxl'))
            return

        archivo = Path(options['archivo'])
        if not archivo.exists():
            self.stderr.write(self.style.ERROR(f'Archivo no encontrado: {archivo}'))
            return

        wb = openpyxl.load_workbook(archivo, read_only=True)
        # Buscar la hoja correcta
        ws = None
        for name in wb.sheetnames:
            if 'REPORTE' in name.upper() or 'IF' in name.upper():
                ws = wb[name]
                break
        if not ws:
            ws = wb.active

        # Agrupar stats por organización política
        stats = {}
        for row in ws.iter_rows(min_row=6, values_only=True):
            op_nombre = row[2] if len(row) > 2 else None
            estado = row[13] if len(row) > 13 else None

            if not op_nombre:
                continue

            op_nombre = str(op_nombre).strip()
            if op_nombre not in stats:
                stats[op_nombre] = {'total': 0, 'presentaron': 0}

            stats[op_nombre]['total'] += 1
            if estado and str(estado).strip() != 'NO PRESENTO':
                stats[op_nombre]['presentaron'] += 1

        wb.close()

        self.stdout.write(f'Encontradas {len(stats)} organizaciones políticas en el Excel\n')

        # Build partido lookup
        partidos = list(Partido.objects.all())
        partido_by_norm_nombre = {}
        partido_by_norm_siglas = {}
        for p in partidos:
            partido_by_norm_nombre[normalize(p.nombre)] = p
            partido_by_norm_siglas[normalize(p.siglas)] = p

        matched = 0
        not_matched = []

        for op_nombre, data in stats.items():
            total = data['total']
            presentaron = data['presentaron']
            porcentaje = round((presentaron / total * 100), 2) if total > 0 else 0

            # Match partido
            norm_op = normalize(op_nombre)
            partido = None

            # 1. Exact match by nombre
            if norm_op in partido_by_norm_nombre:
                partido = partido_by_norm_nombre[norm_op]

            # 2. Check if partido nombre is contained in op_nombre or vice versa
            if not partido:
                for norm_nombre, p in partido_by_norm_nombre.items():
                    if norm_nombre in norm_op or norm_op in norm_nombre:
                        partido = p
                        break

            # 3. Check siglas (op_nombre often has format "NOMBRE - SIGLAS")
            if not partido:
                # Try extracting siglas from the op_nombre
                parts = op_nombre.split('-')
                if len(parts) >= 2:
                    possible_siglas = normalize(parts[-1].strip())
                    if possible_siglas in partido_by_norm_siglas:
                        partido = partido_by_norm_siglas[possible_siglas]

            # 4. Word overlap match
            if not partido:
                STOPWORDS = {'DE', 'DEL', 'LA', 'EL', 'LOS', 'LAS', 'Y', 'EN',
                             'POR', 'PARA', 'PARTIDO', 'POLITICO', 'ALIANZA',
                             'ELECTORAL', 'A', 'MOVIMIENTO', 'NACIONAL'}
                op_words = set(norm_op.split()) - STOPWORDS
                best_score = 0
                for norm_nombre, p in partido_by_norm_nombre.items():
                    p_words = set(norm_nombre.split()) - STOPWORDS
                    common = len(op_words & p_words)
                    if common > best_score and common >= 2:
                        best_score = common
                        partido = p

            if partido:
                partido.transparencia_onpe_total_candidatos = total
                partido.transparencia_onpe_presentaron = presentaron
                partido.transparencia_porcentaje = porcentaje
                partido.transparencia_ultima_actualizacion = date.today()
                partido.save(update_fields=[
                    'transparencia_onpe_total_candidatos',
                    'transparencia_onpe_presentaron',
                    'transparencia_porcentaje',
                    'transparencia_ultima_actualizacion',
                ])
                matched += 1
                self.stdout.write(self.style.SUCCESS(
                    f'  ✓ {op_nombre} → {partido.nombre} ({partido.siglas}) — '
                    f'{presentaron}/{total} ({porcentaje}%)'
                ))
            else:
                not_matched.append(f'{op_nombre} ({total} candidatos)')

        self.stdout.write(self.style.MIGRATE_HEADING(f'\n=== RESUMEN ==='))
        self.stdout.write(self.style.SUCCESS(f'  Partidos actualizados: {matched}/{len(stats)}'))
        if not_matched:
            self.stdout.write(self.style.WARNING(f'  Sin match ({len(not_matched)}):'))
            for nm in not_matched:
                self.stdout.write(self.style.WARNING(f'    ✗ {nm}'))
